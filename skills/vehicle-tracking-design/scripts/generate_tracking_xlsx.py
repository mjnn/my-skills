#!/usr/bin/env python3
"""Generate tracking design xlsx from JSON. Prefers openpyxl for Office-compatible styling."""

from __future__ import annotations

import json
import re
import sys
import zipfile
from pathlib import Path
from xml.sax.saxutils import escape

REQUIRED_PARTS = {
    "[Content_Types].xml",
    "xl/workbook.xml",
    "xl/styles.xml",
    "xl/worksheets/sheet1.xml",
}
MIN_XLSX_BYTES = 8192

# cellXfs style indices
STYLE_DEFAULT = 0
STYLE_HEADER = 1
STYLE_DATA_EVEN = 2
STYLE_DATA_ODD = 3

COL_MIN_WIDTH = 10.0
COL_MAX_WIDTH = 55.0


def col_letter(n: int) -> str:
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def cell_display_width(text: str) -> float:
    """Approximate Excel column width from mixed CJK/Latin content."""
    w = 0.0
    for ch in str(text or ""):
        w += 2.2 if ord(ch) > 127 else 1.0
    return w


def compute_col_widths(rows: list[list[str]]) -> list[float]:
    if not rows:
        return [12.0]
    ncols = max(len(r) for r in rows)
    widths = [COL_MIN_WIDTH] * ncols
    for row in rows:
        for i, val in enumerate(row):
            widths[i] = max(widths[i], cell_display_width(val) * 1.05 + 2)
    return [min(max(w, COL_MIN_WIDTH), COL_MAX_WIDTH) for w in widths]


class XlsxWriter:
    def __init__(self) -> None:
        self.shared_strings: list[str] = []
        self.sheets: list[tuple[str, list[list[str]]]] = []

    def _si(self, text: str) -> int:
        text = "" if text is None else str(text)
        if text not in self.shared_strings:
            self.shared_strings.append(text)
        return self.shared_strings.index(text)

    def add_sheet(self, name: str, rows: list[list[str]]) -> None:
        safe = name[:31].replace("/", "-").replace("\\", "-").replace("*", "-")
        safe = safe.replace("?", "").replace("[", "(").replace("]", ")")
        self.sheets.append((safe or "Sheet", rows))

    def _sheet_xml(self, rows: list[list[str]]) -> str:
        max_col = max((len(r) for r in rows), default=1)
        max_row = len(rows)
        dim_ref = f"A1:{col_letter(max_col)}{max_row}" if rows else "A1"
        col_widths = compute_col_widths(rows)
        filter_ref = f"A1:{col_letter(max_col)}{max_row}" if max_row >= 1 else ""

        # OOXML element order: dimension → sheetViews → sheetFormatPr → cols → sheetData → autoFilter
        parts = [
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
            '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">',
            f'<dimension ref="{dim_ref}"/>',
            "<sheetViews>",
            '<sheetView tabSelected="1" workbookViewId="0">',
            '<selection activeCell="A1" sqref="A1"/>',
            '<pane xSplit="0" ySplit="1" topLeftCell="A2" activePane="bottomLeft" state="frozen"/>',
            "</sheetView>",
            "</sheetViews>",
            '<sheetFormatPr defaultRowHeight="15"/>',
            "<cols>",
        ]
        for i, width in enumerate(col_widths, start=1):
            parts.append(
                f'<col min="{i}" max="{i}" width="{width:.2f}" customWidth="1"/>'
            )
        parts.append("</cols><sheetData>")

        for r_idx, row in enumerate(rows, start=1):
            parts.append(f'<row r="{r_idx}">')
            style = STYLE_HEADER if r_idx == 1 else (
                STYLE_DATA_EVEN if r_idx % 2 == 0 else STYLE_DATA_ODD
            )
            for c_idx, val in enumerate(row, start=1):
                ref = f"{col_letter(c_idx)}{r_idx}"
                si = self._si(val)
                parts.append(f'<c r="{ref}" t="s" s="{style}"><v>{si}</v></c>')
            parts.append("</row>")
        parts.append("</sheetData>")
        if filter_ref:
            parts.append(f'<autoFilter ref="{filter_ref}"/>')
        parts.append("</worksheet>")
        return "".join(parts)

    def _shared_strings_xml(self) -> str:
        return (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            f'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            f'count="{len(self.shared_strings)}" uniqueCount="{len(self.shared_strings)}">'
            + "".join(f"<si><t>{escape(s)}</t></si>" for s in self.shared_strings)
            + "</sst>"
        )

    @staticmethod
    def _styles_xml() -> str:
        return (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            # fonts: 0 normal, 1 header bold white
            '<fonts count="2">'
            '<font><sz val="11"/><color theme="1"/><name val="微软雅黑"/>'
            '<family val="2"/><scheme val="minor"/></font>'
            '<font><b/><sz val="11"/><color rgb="FFFFFFFF"/><name val="微软雅黑"/>'
            '<family val="2"/><scheme val="minor"/></font>'
            "</fonts>"
            # fills: 0 none, 1 gray125, 2 header blue, 3 zebra gray, 4 white
            '<fills count="5">'
            '<fill><patternFill patternType="none"/></fill>'
            '<fill><patternFill patternType="gray125"/></fill>'
            '<fill><patternFill patternType="solid">'
            '<fgColor rgb="FF2F5496"/><bgColor indexed="64"/></patternFill></fill>'
            '<fill><patternFill patternType="solid">'
            '<fgColor rgb="FFF2F2F2"/><bgColor indexed="64"/></patternFill></fill>'
            '<fill><patternFill patternType="solid">'
            '<fgColor rgb="FFFFFFFF"/><bgColor indexed="64"/></patternFill></fill>'
            "</fills>"
            # borders: 0 none, 1 thin grid
            '<borders count="2">'
            "<border><left/><right/><top/><bottom/><diagonal/></border>"
            "<border>"
            '<left style="thin"><color rgb="FFD0D0D0"/></left>'
            '<right style="thin"><color rgb="FFD0D0D0"/></right>'
            '<top style="thin"><color rgb="FFD0D0D0"/></top>'
            '<bottom style="thin"><color rgb="FFD0D0D0"/></bottom>'
            "<diagonal/>"
            "</border>"
            "</borders>"
            '<cellStyleXfs count="1">'
            '<xf numFmtId="0" fontId="0" fillId="0" borderId="0"/>'
            "</cellStyleXfs>"
            '<cellXfs count="4">'
            # 0 default
            '<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>'
            # 1 header
            '<xf numFmtId="0" fontId="1" fillId="2" borderId="1" xfId="0" applyFont="1" '
            'applyFill="1" applyBorder="1" applyAlignment="1">'
            '<alignment horizontal="center" vertical="center" wrapText="1"/>'
            "</xf>"
            # 2 data even (white)
            '<xf numFmtId="0" fontId="0" fillId="4" borderId="1" xfId="0" applyFill="1" '
            'applyBorder="1" applyAlignment="1">'
            '<alignment horizontal="left" vertical="top" wrapText="1"/>'
            "</xf>"
            # 3 data odd (zebra)
            '<xf numFmtId="0" fontId="0" fillId="3" borderId="1" xfId="0" applyFill="1" '
            'applyBorder="1" applyAlignment="1">'
            '<alignment horizontal="left" vertical="top" wrapText="1"/>'
            "</xf>"
            "</cellXfs>"
            '<cellStyles count="1">'
            '<cellStyle name="Normal" xfId="0" builtinId="0"/>'
            "</cellStyles>"
            "</styleSheet>"
        )

    def save(self, path: Path) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        sheet_xmls = [self._sheet_xml(rows) for _, rows in self.sheets]
        n = len(self.sheets)
        styles_rid = n + 1
        sst_rid = n + 2

        with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(
                "[Content_Types].xml",
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
                '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
                '<Default Extension="xml" ContentType="application/xml"/>'
                '<Override PartName="/xl/workbook.xml" '
                'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
                '<Override PartName="/xl/worksheets/sheet1.xml" '
                'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
                + "".join(
                    f'<Override PartName="/xl/worksheets/sheet{i}.xml" '
                    f'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
                    for i in range(2, n + 1)
                )
                + '<Override PartName="/xl/styles.xml" '
                'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>'
                '<Override PartName="/xl/sharedStrings.xml" '
                'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>'
                "</Types>",
            )
            zf.writestr(
                "_rels/.rels",
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                '<Relationship Id="rId1" '
                'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
                'Target="xl/workbook.xml"/>'
                "</Relationships>",
            )
            zf.writestr(
                "xl/_rels/workbook.xml.rels",
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                + "".join(
                    f'<Relationship Id="rId{i}" '
                    f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
                    f'Target="worksheets/sheet{i}.xml"/>'
                    for i in range(1, n + 1)
                )
                + f'<Relationship Id="rId{styles_rid}" '
                'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" '
                'Target="styles.xml"/>'
                + f'<Relationship Id="rId{sst_rid}" '
                'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" '
                'Target="sharedStrings.xml"/>'
                "</Relationships>",
            )
            zf.writestr(
                "xl/workbook.xml",
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
                'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
                "<sheets>"
                + "".join(
                    f'<sheet name="{escape(name)}" sheetId="{i}" r:id="rId{i}"/>'
                    for i, (name, _) in enumerate(self.sheets, start=1)
                )
                + "</sheets></workbook>",
            )
            zf.writestr("xl/styles.xml", self._styles_xml())
            zf.writestr("xl/sharedStrings.xml", self._shared_strings_xml())
            for i, xml in enumerate(sheet_xmls, start=1):
                zf.writestr(f"xl/worksheets/sheet{i}.xml", xml)


def is_nested_event(ev: dict) -> bool:
    return isinstance(ev.get("properties"), list)


def props_to_flat(ev: dict) -> tuple[str, str]:
    props = ev.get("properties") or []
    names, descs = [], []
    for p in props:
        names.append(str(p.get("prop_name", "")))
        cn = p.get("prop_chinese_name") or p.get("prop_name") or ""
        descs.append(str(cn))
    return ",".join(names), ",".join(descs)


def normalize_events(events: list[dict]) -> list[list[str]]:
    header = ["事件ID", "事件名称", "事件说明", "分析目的", "核心指标", "上报参数", "参数说明", "数据分类"]
    rows = [header]
    for ev in events:
        if is_nested_event(ev):
            params, params_desc = props_to_flat(ev)
            rows.append([
                ev.get("event_id", ""),
                ev.get("event_chinese_name") or ev.get("event_name", ""),
                ev.get("trigger_condition", ""),
                ev.get("analysis_purpose", ""),
                ev.get("core_metric", ""),
                params,
                params_desc,
                ev.get("data_category", "车辆使用数据"),
            ])
        else:
            rows.append([
                ev.get("event_id", ""),
                ev.get("event_name", ""),
                ev.get("event_desc", ""),
                ev.get("analysis_purpose", ""),
                ev.get("core_metric", ""),
                ev.get("params", ""),
                ev.get("params_desc", ""),
                ev.get("data_category", "车辆使用数据"),
            ])
    return rows


def normalize_paths(paths: list[dict]) -> list[list[str]]:
    header = ["路径ID", "路径名称", "路径描述", "关键节点", "转化目标", "埋点事件", "分析指标", "数据分类"]
    rows = [header]
    for p in paths:
        if isinstance(p.get("steps"), list):
            nodes = " → ".join(
                s.get("description") or s.get("event_name") or s.get("event_id", "")
                for s in sorted(p["steps"], key=lambda x: x.get("step_order", 0))
            )
            events = ",".join(s.get("event_id", "") for s in p["steps"])
            rows.append([
                p.get("path_id", ""),
                p.get("path_name", ""),
                p.get("path_description", ""),
                nodes,
                p.get("conversion_target", ""),
                events,
                p.get("core_indicator", ""),
                p.get("data_category", "车辆使用数据"),
            ])
        else:
            rows.append([
                p.get("path_id", ""),
                p.get("path_name", ""),
                p.get("path_desc", ""),
                p.get("key_nodes", ""),
                p.get("conversion_target", ""),
                p.get("tracking_events", ""),
                p.get("analysis_metric", ""),
                p.get("data_category", "车辆使用数据"),
            ])
    return rows


def normalize_funnels(funnels: list[dict]) -> list[list[str]]:
    header = ["漏斗ID", "漏斗名称", "漏斗步骤", "步骤描述", "埋点事件", "转化指标", "优化目标", "数据分类"]
    rows = [header]
    for f in funnels:
        if isinstance(f.get("steps"), list):
            for step in sorted(f["steps"], key=lambda x: x.get("step_order", 0)):
                rows.append([
                    f.get("funnel_id", ""),
                    f.get("funnel_name", ""),
                    f"Step{step.get('step_order', '')}: {step.get('step_name', '')}",
                    step.get("description", ""),
                    step.get("event_id", ""),
                    f.get("primary_metric", ""),
                    f.get("optimization_target", ""),
                    f.get("data_category", "车辆使用数据"),
                ])
        else:
            rows.append([
                f.get("funnel_id", ""),
                f.get("funnel_name", ""),
                f.get("funnel_steps", ""),
                f.get("step_desc", ""),
                f.get("tracking_events", ""),
                f.get("conversion_metric", ""),
                f.get("optimization_target", ""),
                f.get("data_category", "车辆使用数据"),
            ])
    return rows


def normalize_relations(relations: list[dict]) -> list[list[str]]:
    header = ["关联ID", "关联类型", "源事件", "目标事件", "关联关系", "埋点参数", "分析目的", "数据分类"]
    rows = [header]
    for r in relations:
        if "relation_id" in r or "from_event" in r:
            rows.append([
                r.get("relation_id", ""),
                r.get("relation_type", ""),
                r.get("from_event", ""),
                r.get("to_event", ""),
                r.get("description", ""),
                r.get("params", ""),
                r.get("analysis_purpose", ""),
                r.get("data_category", "车辆使用数据"),
            ])
        else:
            rows.append([
                r.get("rel_id", ""),
                r.get("rel_type", ""),
                r.get("source_event", ""),
                r.get("target_event", ""),
                r.get("relation_desc", ""),
                r.get("params", ""),
                r.get("analysis_purpose", ""),
                r.get("data_category", "车辆使用数据"),
            ])
    return rows


def validate_xlsx(path: Path, min_event_rows: int = 1) -> list[str]:
    """Return validation error messages; empty list means OK."""
    errors: list[str] = []
    if not path.is_file():
        return [f"file not found: {path}"]
    size = path.stat().st_size
    if size < MIN_XLSX_BYTES:
        errors.append(f"file too small ({size} bytes); likely empty sharedStrings")

    try:
        with zipfile.ZipFile(path) as zf:
            names = set(zf.namelist())
            missing = REQUIRED_PARTS - names
            if missing:
                errors.append(f"missing OOXML parts: {sorted(missing)}")

            sst_name = "xl/sharedStrings.xml"
            if sst_name in names:
                sst = zf.read(sst_name).decode("utf-8")
                m = re.search(r'uniqueCount="(\d+)"', sst)
                unique = int(m.group(1)) if m else 0
                if unique == 0:
                    errors.append("sharedStrings uniqueCount is 0")
                sh1 = zf.read("xl/worksheets/sheet1.xml").decode("utf-8")
                indices = [int(v) for v in re.findall(r't="s"><v>(\d+)</v>', sh1)]
                if indices and max(indices) >= unique:
                    errors.append(
                        f"string index out of range: max={max(indices)}, uniqueCount={unique}"
                    )
            else:
                sh1 = zf.read("xl/worksheets/sheet1.xml").decode("utf-8")
                if "<v>" not in sh1 and "<is>" not in sh1:
                    errors.append("sheet1 has no cell values")
            rows = sh1.count("<row ")
            expected = min_event_rows + 1
            if rows < expected:
                errors.append(f"sheet1 has {rows} rows, expected at least {expected}")
    except zipfile.BadZipFile:
        errors.append("not a valid zip/xlsx file")

    if not errors:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            if not wb.sheetnames:
                errors.append("openpyxl: workbook has no sheets")
            elif wb[wb.sheetnames[0]].max_row is None or wb[wb.sheetnames[0]].max_row < 2:
                errors.append("openpyxl: first sheet appears empty")
            wb.close()
        except ImportError:
            pass
        except Exception as exc:
            errors.append(f"openpyxl cannot open file: {exc}")

    return errors


def _write_styled_openpyxl(sheets: list[tuple[str, list[list[str]]]], output: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="2F5496")
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="微软雅黑", size=11)
    zebra_fill = PatternFill("solid", fgColor="F2F2F2")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(wrap_text=True, horizontal="center", vertical="center")

    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in sheets:
        ws = wb.create_sheet(title=name[:31])
        if not rows:
            continue
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = border
                if r_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center
                else:
                    cell.fill = zebra_fill if r_idx % 2 == 0 else white_fill
                    cell.font = data_font
                    cell.alignment = wrap
        for i, width in enumerate(compute_col_widths(rows), start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
        ws.freeze_panes = "A2"
        if rows:
            max_col = max(len(r) for r in rows)
            ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{len(rows)}"
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def generate(data: dict, output: Path) -> list[str]:
    app_name = data.get("app_name", "APP")
    events = data.get("events", [])
    sheets = [
        (f"{app_name}埋点设计", normalize_events(events)),
        ("用户路径埋点", normalize_paths(data.get("paths", []))),
        ("转化漏斗埋点", normalize_funnels(data.get("funnels", []))),
        ("关联埋点设计", normalize_relations(data.get("relations", []))),
    ]
    try:
        _write_styled_openpyxl(sheets, output)
    except ImportError:
        writer = XlsxWriter()
        for name, rows in sheets:
            writer.add_sheet(name, rows)
        writer.save(output)
    return validate_xlsx(output, min_event_rows=max(len(events), 1))


def main() -> int:
    argv = sys.argv[1:]
    if len(argv) == 2 and argv[0] == "--validate-only":
        errors = validate_xlsx(Path(argv[1]))
        if errors:
            print("VALIDATION FAILED:", file=sys.stderr)
            for e in errors:
                print(f"  - {e}", file=sys.stderr)
            return 1
        print(f"OK: {argv[1]}")
        return 0

    if len(argv) != 2:
        print(
            "Usage:\n"
            "  python generate_tracking_xlsx.py <input.json> <output.xlsx>\n"
            "  python generate_tracking_xlsx.py --validate-only <output.xlsx>",
            file=sys.stderr,
        )
        return 1

    input_path = Path(argv[0])
    output_path = Path(argv[1])
    with input_path.open(encoding="utf-8") as f:
        data = json.load(f)
    errors = generate(data, output_path)
    if errors:
        print(f"Generated but VALIDATION FAILED: {output_path}", file=sys.stderr)
        for e in errors:
            print(f"  - {e}", file=sys.stderr)
        return 1
    print(f"Generated: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
